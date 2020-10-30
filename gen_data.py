import openpyxl
import uuid

from datetime import datetime
from random import choice, randint, random
from faker import Faker
fake = Faker()


def date_to_string(value):
    return datetime.strftime(value, '%Y-%m-%d')


def create_workbook(filename):
    wb = openpyxl.Workbook()
    wb.save(filename)


def gen_categoria(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet']

    ws['A1'] = 'id'
    ws['B1'] = 'categoria'

    max_row = ws.max_row
    categories = [
        'bebidas',
        'sobremesas',
        'cereais',
        'frios',
        'limpeza',
        'hortifruti',
        'perfumaria',
        'mercearia',
        'bazar',
        'pet shop',
    ]

    ids = []
    for i, row in enumerate(categories, 2):
        _id = randint(100, 1000)
        ids.append(_id)
        ws[f'A{i}'] = _id
        ws[f'B{i}'] = row
        print(i, row)
    wb.save(filename)
    return ids


def gen_produto(filename, categorias_id):
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet']

    ws['A1'] = 'id'
    ws['B1'] = 'produto'
    ws['C1'] = 'preco'
    ws['D1'] = 'categoria_id'

    max_row = ws.max_row
    number_rows = 50

    ids_e_precos = []
    for i, row in enumerate(range(max_row + 1, max_row + number_rows + 1)):
        _id = randint(100, 1000)
        price = round(randint(100, 1000) * random(), 2)
        # Gera uma tupla
        ids_e_precos.append((_id, price))
        product = ' '.join(fake.words(nb=2))
        print(product)
        ws[f'A{row}'] = _id
        ws[f'B{row}'] = product
        ws[f'C{row}'] = price
        ws[f'D{row}'] = choice(categorias_id)
    wb.save(filename)
    return ids_e_precos


def gen_cliente(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet']

    ws['A1'] = 'id'
    ws['B1'] = 'nome'

    max_row = ws.max_row

    ids = []
    for i, row in enumerate(range(50), 2):
        _id = randint(100, 1000)
        ids.append(_id)
        ws[f'A{i}'] = _id
        ws[f'B{i}'] = fake.name()
    wb.save(filename)
    return ids


def gen_venda(filename, clientes_id):
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet']

    ws['A1'] = 'id'
    ws['B1'] = 'data'
    ws['C1'] = 'cliente_id'
    ws['D1'] = 'slug'

    max_row = ws.max_row
    number_rows = 100

    ids = []
    for i, row in enumerate(range(max_row + 1, max_row + number_rows + 1)):
        _id = randint(100, 1000)
        ids.append(_id)
        data = datetime(2020, randint(1, 12), randint(1, 28))
        ws[f'A{row}'] = _id
        ws[f'B{row}'] = date_to_string(data)
        ws[f'C{row}'] = choice(clientes_id)
        ws[f'D{row}'] = str(uuid.uuid4())
    wb.save(filename)
    return ids


def gen_itens_venda(filename, vendas_id, produtos_id):
    wb = openpyxl.load_workbook(filename)
    ws = wb['Sheet']

    ws['A1'] = 'id'
    ws['B1'] = 'venda_id'
    ws['C1'] = 'produto_id'
    ws['D1'] = 'quantidade'
    ws['E1'] = 'preco_venda'

    max_row = ws.max_row
    number_rows = 2000

    for i, row in enumerate(range(max_row + 1, max_row + number_rows + 1)):
        produto = choice(produtos_id)
        ws[f'A{row}'] = i + 1
        ws[f'B{row}'] = choice(vendas_id)
        ws[f'C{row}'] = produto[0]
        ws[f'D{row}'] = randint(1, 100)
        ws[f'E{row}'] = produto[1]
    wb.save(filename)


if __name__ == '__main__':
    filename = 'categoria.xlsx'
    create_workbook(filename)
    categorias_id = gen_categoria(filename)

    filename = 'produto.xlsx'
    create_workbook(filename)
    produtos_id = gen_produto(filename, categorias_id)

    filename = 'cliente.xlsx'
    create_workbook(filename)
    clientes_id = gen_cliente(filename)

    filename = 'venda.xlsx'
    create_workbook(filename)
    vendas_id = gen_venda(filename, clientes_id)

    filename = 'itens_venda.xlsx'
    create_workbook(filename)
    gen_itens_venda(filename, vendas_id, produtos_id)
